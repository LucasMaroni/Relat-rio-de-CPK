#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador do Relatório Analítico de Custo por Quilômetro — Grupo Maroni.

Como funciona:
  1. Procura, na pasta ./dados, as DUAS planilhas do mês:
        - o RELATÓRIO de CPK por placa  (detectado pela coluna CPK / KM ATUAL)
        - a BASE OFICIAL de notas fiscais (detectada pela coluna 'DAT. REFER')
     Os nomes dos arquivos NÃO importam — basta serem .xlsx e estarem na pasta ./dados.
  2. Recalcula todos os indicadores e a série mensal.
  3. Injeta os dados no template.html e grava o index.html publicado.

Rodar manualmente:  python gerar_relatorio.py
No GitHub:          roda sozinho a cada envio de planilha (ver .github/workflows/deploy.yml)
"""
import os, re, json, glob, sys
import pandas as pd
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DADOS_DIR = os.path.join(BASE_DIR, "dados")
TEMPLATE = os.path.join(BASE_DIR, "template.html")
SAIDA = os.path.join(BASE_DIR, "index.html")

# Mapeamento de agrupamento de modelos (aprovado pelo gestor).
# Para unir novos modelos escritos de formas diferentes, acrescente linhas aqui:
#   'NOME COMO ESTÁ NA PLANILHA (maiúsculo)': 'NOME AGRUPADO',
OVERRIDE = {
    'MBATEGO 2429 - SIDER': 'MB ATEGO 2429',
    'SCANIA RH 450 PLUS':  'SCANIA R450 PLUS',
    'SCANIA RN 450 PLUS':  'SCANIA R450 PLUS',
    'SCANIA RH 500 NA':    'SCANIA R500 NA',
    'SCANIA R500 NA':      'SCANIA R500 NA',
}
MES_MINIMO = '2023-01'   # a curva mensal começa a partir daqui


def _colapsa(s):
    return re.sub(r'\s+', ' ', str(s).upper().strip())


def _detecta_planilhas():
    arquivos = sorted(glob.glob(os.path.join(DADOS_DIR, "*.xlsx")))
    arquivos = [a for a in arquivos if not os.path.basename(a).startswith("~$")]
    if len(arquivos) < 2:
        sys.exit("ERRO: coloque as DUAS planilhas (.xlsx) na pasta 'dados/'. "
                 f"Encontrei {len(arquivos)}.")
    relatorio = base = None
    for a in arquivos:
        wb = openpyxl.load_workbook(a, read_only=True, data_only=True)
        achou_base = False
        for ws in wb.worksheets:
            cab = [str(c.value).strip().upper() if c.value is not None else ''
                   for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if any('DAT. REFER' in c or c == 'DAT. REFER' for c in cab) or 'VR. LIQ' in cab:
                achou_base = True; break
        wb.close()
        if achou_base and base is None:
            base = a
        elif relatorio is None:
            relatorio = a
    if not relatorio or not base:
        # fallback: assume o maior arquivo = base
        arquivos.sort(key=os.path.getsize)
        base = base or arquivos[-1]
        relatorio = relatorio or arquivos[0]
    print(f"  Relatório de CPK : {os.path.basename(relatorio)}")
    print(f"  Base de notas    : {os.path.basename(base)}")
    return relatorio, base


def _acha_aba_relatorio(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    for ws in wb.worksheets:
        for r in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            vals = [str(v).strip().upper() if v is not None else '' for v in r]
            if 'PLACA' in vals and 'CPK' in vals and 'CUSTO DE MANUTENÇÃO' in vals:
                wb.close(); return ws.title, r, vals
    wb.close()
    return wb.worksheets[0].title if wb.worksheets else 0, None, None


def carrega_relatorio(caminho):
    # Descobre em qual linha está o cabeçalho (MARCA/PLACA/...).
    raw = pd.read_excel(caminho, sheet_name=0, header=None)
    hdr = None
    for i in range(min(8, len(raw))):
        linha = [str(x).strip().upper() for x in raw.iloc[i].tolist()]
        if 'PLACA' in linha and 'CPK' in linha:
            hdr = i; break
    if hdr is None:
        sys.exit("ERRO: não encontrei o cabeçalho (MARCA/PLACA/...) no relatório de CPK.")
    df = pd.read_excel(caminho, sheet_name=0, skiprows=hdr)
    df.columns = [str(c).strip().upper() for c in df.columns]
    # o relatório tem um bloco-resumo lateral que repete 'MARCA' etc.: mantém só a 1ª ocorrência
    df = df.loc[:, ~pd.Index(df.columns).duplicated()]
    ren = {'CUSTO DE MANUTENÇÃO': 'CUSTO', 'KM ATUAL': 'KM', 'OPERAÇÃO': 'OPERACAO'}
    df = df.rename(columns=ren)
    need = ['MARCA', 'PLACA', 'MODELO', 'TIPO', 'OPERACAO', 'KM', 'ANO', 'CHASSI', 'CUSTO', 'CPK']
    df = df[[c for c in need if c in df.columns]].copy()
    df = df.dropna(subset=['PLACA'])
    for c in ['MARCA', 'PLACA', 'MODELO', 'TIPO', 'OPERACAO', 'CHASSI', 'ANO']:
        if c in df: df[c] = df[c].astype(str).str.strip()
    df['KM'] = pd.to_numeric(df['KM'], errors='coerce').fillna(0)
    df['CUSTO'] = pd.to_numeric(df['CUSTO'], errors='coerce').fillna(0)
    df['CPK'] = pd.to_numeric(df['CPK'], errors='coerce').fillna(0)
    df = df[df['PLACA'].str.upper() != 'NAN']
    df['MO'] = df['MODELO'].map(lambda m: OVERRIDE.get(_colapsa(m), _colapsa(m)))
    return df


def carrega_base(caminho, placas_norm, tot_km):
    # Descobre a aba e as colunas de data/valor/placa.
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    aba = None
    for ws in wb.worksheets:
        cab = [str(c.value).strip().upper() if c.value is not None else ''
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if any('DAT. REFER' in c for c in cab) or 'VR. LIQ' in cab:
            aba = ws.title; break
    wb.close()
    if aba is None:
        print("  AVISO: base de notas sem colunas reconhecidas — série mensal vazia.")
        return [], 0.0
    df = pd.read_excel(caminho, sheet_name=aba)
    df.columns = [str(c).strip().upper() for c in df.columns]
    col_data = next((c for c in df.columns if c.startswith('DAT. REFER') or c == 'DAT. REFER'), None)
    col_val = 'VR. LIQ' if 'VR. LIQ' in df.columns else next((c for c in df.columns if 'LIQ' in c), None)
    col_placa = 'PLACA' if 'PLACA' in df.columns else next((c for c in df.columns if c == 'PLACA'), None)
    if not (col_data and col_val and col_placa):
        print("  AVISO: colunas DAT. REFER / VR. LIQ / PLACA não encontradas — série mensal vazia.")
        return [], 0.0
    df['K'] = df[col_placa].map(lambda s: re.sub(r'[^A-Z0-9]', '', str(s).upper()))
    df['DT'] = pd.to_datetime(df[col_data], errors='coerce')
    df = df[df['K'].isin(placas_norm) & df['DT'].notna()].copy()
    df['ym'] = df['DT'].dt.strftime('%Y-%m')
    g = df.groupby('ym')[col_val].sum().reset_index().sort_values('ym')
    g = g[g['ym'] >= MES_MINIMO]
    cobertura = float(df[col_val].sum())
    mensal = []
    for _, row in g.iterrows():
        cu = float(row[col_val])
        mensal.append({'ym': row['ym'], 'cu': round(cu, 2), 'cpk': round(cu / tot_km, 6)})
    return mensal, cobertura


def main():
    print("Gerando relatório...")
    if not os.path.isdir(DADOS_DIR):
        sys.exit("ERRO: crie a pasta 'dados/' e coloque as duas planilhas nela.")
    rel_path, base_path = _detecta_planilhas()

    rep = carrega_relatorio(rel_path)
    tot_c = float(rep['CUSTO'].sum())
    tot_k = float(rep['KM'].sum())
    placas_norm = set(rep['PLACA'].map(lambda s: re.sub(r'[^A-Z0-9]', '', str(s).upper())))

    rows = [{'ma': r.MARCA, 'pl': r.PLACA, 'mo': r.MO, 'mo0': r.MODELO, 'ti': r.TIPO,
             'op': r.OPERACAO, 'km': int(r.KM), 'an': str(r.ANO), 'ch': str(r.CHASSI),
             'cu': round(float(r.CUSTO), 2), 'cpk': round(float(r.CPK), 6)}
            for r in rep.itertuples(index=False)]

    mensal, cobertura = carrega_base(base_path, placas_norm, tot_k)
    meta = {'totC': round(tot_c, 2), 'totK': int(tot_k), 'baseCov': round(cobertura, 2)}

    with open(TEMPLATE, encoding='utf-8') as f:
        html = f.read()
    html = html.replace('/*__DATA__*/', json.dumps(rows, ensure_ascii=False, separators=(',', ':')))
    html = html.replace('/*__MONTHLY__*/', json.dumps(mensal, ensure_ascii=False, separators=(',', ':')))
    html = html.replace('/*__META__*/', json.dumps(meta, ensure_ascii=False))
    with open(SAIDA, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"  Veículos: {len(rows)} | Custo total: R$ {tot_c:,.2f} | Meses: {len(mensal)}")
    print(f"  index.html gerado com sucesso ({os.path.getsize(SAIDA)//1024} KB).")


if __name__ == '__main__':
    main()
